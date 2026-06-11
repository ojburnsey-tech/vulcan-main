# api/measurement_import.py
# Phase-1 measurement import: parse an uploaded CSV or XLSX into plain rows of
# {description, quantity, unit}. Deliberately dumb — no AI, no pricing, no NRM2,
# no rate matching, no persistence. It only reads the file and hands back rows.
#
# Kept in its own module (like export_pdf.py / export_excel.py) so it is fully
# isolated from the /process pipeline and easy to test on its own.

import csv
import io
import re


class MeasurementImportError(Exception):
    """A user-facing parsing failure. `status` is the HTTP code the route returns."""

    def __init__(self, message, status=422):
        super().__init__(message)
        self.message = message
        self.status = status


# Cap parsed rows so a pathological file can't balloon the response.
_MAX_ROWS = 5000

# Header synonyms (lower-cased) used to locate the three columns we care about.
# Matched by exact membership first, then by the prefix/substring rules in
# _classify_header below, so reordered or slightly-renamed columns still map.
_DESC_HEADERS = {'description', 'desc', 'item', 'measurement', 'label', 'name',
                 'subject', 'comment', 'comments', 'space'}
_QTY_HEADERS  = {'quantity', 'qty', 'amount', 'value', 'measure', 'count',
                 'length', 'area', 'volume', 'total'}
_UNIT_HEADERS = {'unit', 'units', 'uom', 'unit of measure', 'measurement unit'}


def _text(value):
    """Trimmed string form of a cell value; '' for None."""
    if value is None:
        return ''
    return str(value).strip()


def _norm(value):
    """Lower-cased trimmed form, for header matching."""
    return _text(value).lower()


def _parse_quantity(raw):
    """Best-effort numeric quantity. Returns int/float, or None when absent.

    Accepts native numbers (from XLSX) and strings like '12', '12.5', '1,250',
    or '45 m3' (first number wins). Non-numeric / empty cells return None so the
    caller can decide what to do — we never guess a quantity.
    """
    if raw is None or raw == '':
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        num = float(raw)
    else:
        m = re.search(r'-?\d+(?:\.\d+)?', str(raw).replace(',', ''))
        if not m:
            return None
        num = float(m.group())
    return int(num) if num.is_integer() else num


def _classify_header(header_cell):
    """Return 'description' | 'quantity' | 'unit' | None for one header cell."""
    h = _norm(header_cell)
    if not h:
        return None
    if h in _DESC_HEADERS or h.startswith('desc'):
        return 'description'
    if h in _UNIT_HEADERS or h.startswith('unit') or h == 'uom':
        return 'unit'
    if h in _QTY_HEADERS or h.startswith('qty') or 'quantit' in h:
        return 'quantity'
    return None


def _map_columns(header_row):
    """Map recognised header cells to column indices. Empty dict if none match
    (caller then falls back to positional columns)."""
    mapping = {}
    for idx, cell in enumerate(header_row):
        field = _classify_header(cell)
        if field and field not in mapping:
            mapping[field] = idx
    return mapping


def _rows_to_measurements(rows):
    """Convert a list of raw cell-rows into measurement dicts.

    Drops fully-empty rows, detects an optional header row, and keeps only rows
    that have a description. Raises MeasurementImportError when nothing usable
    remains.
    """
    cleaned = [list(r) for r in rows if r is not None and any(_text(c) for c in r)]
    if not cleaned:
        raise MeasurementImportError("The file contains no data rows.", 422)

    col = _map_columns(cleaned[0])
    if col:                                   # first row is a recognised header
        data_rows = cleaned[1:]
        di, qi, ui = col.get('description'), col.get('quantity'), col.get('unit')
    else:                                     # no header → assume desc | qty | unit
        data_rows = cleaned
        di, qi, ui = 0, 1, 2

    def cell(row, i):
        return row[i] if (i is not None and i < len(row)) else None

    measurements = []
    for row in data_rows:
        description = _text(cell(row, di))
        if not description:
            continue                          # a measurement must have a description
        measurements.append({
            "description": description,
            "quantity":    _parse_quantity(cell(row, qi)),
            "unit":        _text(cell(row, ui)),
        })
        if len(measurements) >= _MAX_ROWS:
            break

    if not measurements:
        raise MeasurementImportError(
            "No measurements with a description were found. Expected columns: "
            "description, quantity, unit.", 422)
    return measurements


def _parse_csv(file_bytes):
    """Decode and read a CSV into raw rows; sniffs the delimiter."""
    text = None
    for encoding in ('utf-8-sig', 'latin-1'):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise MeasurementImportError("Could not decode the CSV file as text.", 422)

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|") if sample.strip() else csv.excel
    except csv.Error:
        dialect = csv.excel                   # fall back to comma-separated

    try:
        rows = list(csv.reader(io.StringIO(text), dialect))
    except csv.Error as exc:
        raise MeasurementImportError(f"The CSV file is malformed: {exc}", 422)

    if not rows:
        raise MeasurementImportError("The CSV file is empty.", 422)
    return rows


def _parse_xlsx(file_bytes):
    """Read the active sheet of an XLSX into raw rows."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise MeasurementImportError("XLSX support is not available on the server.", 500)

    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        # openpyxl raises a variety of types (BadZipFile, InvalidFileException, …)
        # for corrupt or non-workbook input — all map to the same user message.
        raise MeasurementImportError(
            "Could not read the XLSX file. It may be corrupt or not a valid "
            "Excel workbook.", 422)

    try:
        ws = wb.active
        if ws is None:
            raise MeasurementImportError("The workbook has no active sheet.", 422)
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    if not rows:
        raise MeasurementImportError("The XLSX file is empty.", 422)
    return rows


def parse_measurements(filename, file_bytes):
    """Parse an uploaded measurement file into a list of {description, quantity, unit}.

    Args:
        filename:   original upload name — only its extension is used to route.
        file_bytes: raw file contents.

    Returns:
        list[dict] with keys description (str), quantity (int/float/None), unit (str).

    Raises:
        MeasurementImportError with a user-facing .message and HTTP .status.
    """
    name = (filename or '').lower()
    if not file_bytes:
        raise MeasurementImportError("The uploaded file is empty.", 400)

    if name.endswith('.csv'):
        rows = _parse_csv(file_bytes)
    elif name.endswith('.xlsx'):
        rows = _parse_xlsx(file_bytes)
    else:
        raise MeasurementImportError(
            "Unsupported file type. Upload a .csv or .xlsx file.", 415)

    return _rows_to_measurements(rows)
