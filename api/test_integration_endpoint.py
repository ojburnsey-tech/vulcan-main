# api/test_integration_endpoint.py
# Integration tests for the measurement classification endpoint.
#
# Tests:
# - Endpoint responds to POST requests
# - Endpoint returns correct classification for known measurements
# - Endpoint returns {matched: false} for unknown measurements
# - Endpoint handles invalid input gracefully
# - Endpoint returns proper JSON structure

import unittest
import json
import tempfile
from pathlib import Path
from openpyxl import Workbook
from mapping_loader import _loader
import app as app_module


class TestMeasurementLookupEndpoint(unittest.TestCase):
    """Integration tests for /measurement/lookup-mapping endpoint."""

    @classmethod
    def setUpClass(cls):
        """Set up test Flask app and load mappings."""
        # Create test spreadsheet
        cls.temp_dir = tempfile.mkdtemp()
        cls.test_file = Path(cls.temp_dir) / "Bluebeam_Term_Mapping_1.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Mappings"

        headers = [
            "Measurement Description",
            "Trade Code",
            "Trade Group",
            "CSI Division",
            "Unit",
            "Takeoff Type",
        ]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        test_data = [
            ["Ceramic floor tile", "TILE-CERAMIC", "Finishes", "09 3000", "m²", "Flooring"],
            ["Brick cavity wall", "BRICK-CAV", "Masonry", "04 4000", "m²", "Brickwork"],
            ["Reinforced concrete foundations", "CONC-FDN", "Structural", "03 3000", "m³", "Concreting"],
            ["Roof tiles pitched roof", "TILE-PITCH", "Roofing", "07 7000", "m²", "Roofing"],
            ["Emulsion paint to walls", "PAINT-EMULSION", "Finishes", "09 9100", "m²", "Decoration"],
            ["Wall Area", "WALL-AREA", "Walls", "04 4000", "m²", "Measurement"],
        ]

        for row_num, row_data in enumerate(test_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        wb.save(cls.test_file)

        # Load mappings
        _loader._file_path = cls.test_file
        _loader.load()

        # Create Flask test client
        app_module.app.config['TESTING'] = True
        cls.client = app_module.app.test_client()

    @classmethod
    def tearDownClass(cls):
        """Clean up."""
        import shutil
        shutil.rmtree(cls.temp_dir, ignore_errors=True)

    def test_endpoint_exists(self):
        """Test that the endpoint exists and responds."""
        response = self.client.post(
            "/measurement/lookup-mapping",
            data=json.dumps({"description": "Wall Area"}),
            content_type="application/json",
        )
        self.assertIn(response.status_code, [200, 400, 422])

    def test_endpoint_match_found(self):
        """Test endpoint returns classification for known measurement."""
        response = self.client.post(
            "/measurement/lookup-mapping",
            data=json.dumps({"description": "Wall Area"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)

        self.assertTrue(data["matched"])
        self.assertEqual(data["description"], "Wall Area")
        self.assertEqual(data["trade_code"], "WALL-AREA")
        self.assertEqual(data["trade_group"], "Walls")
        self.assertEqual(data["csi_division"], "04 4000")
        self.assertEqual(data["unit"], "m²")
        self.assertEqual(data["takeoff_type"], "Measurement")

    def test_endpoint_case_insensitive(self):
        """Test endpoint is case-insensitive."""
        response = self.client.post(
            "/measurement/lookup-mapping",
            data=json.dumps({"description": "CERAMIC FLOOR TILE"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)

        self.assertTrue(data["matched"])
        self.assertEqual(data["trade_code"], "TILE-CERAMIC")

    def test_endpoint_with_whitespace(self):
        """Test endpoint trims whitespace."""
        response = self.client.post(
            "/measurement/lookup-mapping",
            data=json.dumps({"description": "  Brick cavity wall  "}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)

        self.assertTrue(data["matched"])
        self.assertEqual(data["trade_code"], "BRICK-CAV")
        self.assertEqual(data["description"], "Brick cavity wall")

    def test_endpoint_no_match(self):
        """Test endpoint returns matched=false for unknown measurement."""
        response = self.client.post(
            "/measurement/lookup-mapping",
            data=json.dumps({"description": "Unknown measurement type"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)

        self.assertFalse(data["matched"])
        self.assertEqual(data["description"], "Unknown measurement type")
        self.assertNotIn("trade_code", data)

    def test_endpoint_empty_description(self):
        """Test endpoint with empty description."""
        response = self.client.post(
            "/measurement/lookup-mapping",
            data=json.dumps({"description": ""}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)

        self.assertFalse(data["matched"])

    def test_endpoint_missing_description(self):
        """Test endpoint with missing description field."""
        response = self.client.post(
            "/measurement/lookup-mapping",
            data=json.dumps({}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)

        self.assertFalse(data["matched"])

    def test_endpoint_null_description(self):
        """Test endpoint with null description."""
        response = self.client.post(
            "/measurement/lookup-mapping",
            data=json.dumps({"description": None}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)

        self.assertFalse(data["matched"])

    def test_endpoint_invalid_json(self):
        """Test endpoint with invalid JSON."""
        response = self.client.post(
            "/measurement/lookup-mapping",
            data="not json",
            content_type="application/json",
        )

        # Should either return 200 or error status
        self.assertIn(response.status_code, [200, 400, 422])

    def test_endpoint_multiple_requests(self):
        """Test endpoint can handle multiple requests."""
        test_cases = [
            ("Ceramic floor tile", True, "TILE-CERAMIC"),
            ("Brick cavity wall", True, "BRICK-CAV"),
            ("Unknown", False, None),
            ("Reinforced concrete foundations", True, "CONC-FDN"),
        ]

        for description, should_match, expected_code in test_cases:
            response = self.client.post(
                "/measurement/lookup-mapping",
                data=json.dumps({"description": description}),
                content_type="application/json",
            )

            self.assertEqual(response.status_code, 200)
            data = json.loads(response.data)

            self.assertEqual(data["matched"], should_match)
            if should_match:
                self.assertEqual(data["trade_code"], expected_code)

    def test_endpoint_response_format_matched(self):
        """Test response format for matched case."""
        response = self.client.post(
            "/measurement/lookup-mapping",
            data=json.dumps({"description": "Wall Area"}),
            content_type="application/json",
        )

        data = json.loads(response.data)

        # Must have these fields
        required_fields = [
            "matched",
            "description",
            "trade_code",
            "trade_group",
            "csi_division",
            "unit",
            "takeoff_type",
        ]
        for field in required_fields:
            self.assertIn(field, data, f"Missing field in response: {field}")

    def test_endpoint_response_format_not_matched(self):
        """Test response format for not matched case."""
        response = self.client.post(
            "/measurement/lookup-mapping",
            data=json.dumps({"description": "Unknown"}),
            content_type="application/json",
        )

        data = json.loads(response.data)

        # Must have these fields
        self.assertIn("matched", data)
        self.assertIn("description", data)
        self.assertFalse(data["matched"])


if __name__ == "__main__":
    unittest.main()
