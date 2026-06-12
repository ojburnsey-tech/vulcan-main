# api/mapping_loader.py
# Load and cache measurement mappings from Bluebeam_Term_Mapping_1.xlsx
# Runs once at application startup to prevent repeated disk I/O.
#
# The spreadsheet is the authoritative source for measurement classification.
# No mappings are hardcoded in Python — all data comes from the Excel file.

import os
import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


logger = logging.getLogger(__name__)


class MappingLoader:
    """Loads and caches measurement classification mappings from the Bluebeam
    Term Mapping spreadsheet. Provides fast lookup functions for exact and
    case-insensitive matching.
    """

    # Expected column headers in the spreadsheet
    REQUIRED_COLUMNS = [
        "Measurement Description",
        "Trade Code",
        "Trade Group",
        "CSI Division",
        "Unit",
        "Takeoff Type",
    ]

    def __init__(self):
        """Initialize the loader. Mappings are loaded during first load() call."""
        self._mappings: List[Dict] = []
        self._lookup_index: Dict[str, Dict] = {}  # lowercase description -> mapping
        self._loaded = False
        self._file_path = self._find_mapping_file()

    @staticmethod
    def _find_mapping_file() -> Optional[Path]:
        """Search for Bluebeam_Term_Mapping_1.xlsx in the same directory as this
        file and in the parent directory."""
        candidates = [
            Path(__file__).parent / "Bluebeam_Term_Mapping_1.xlsx",
            Path(__file__).parent.parent / "Bluebeam_Term_Mapping_1.xlsx",
        ]
        for path in candidates:
            if path.exists():
                return path
        return None

    def _validate_columns(self, headers: List[str]) -> Tuple[bool, Optional[str]]:
        """Verify that all required columns exist in the spreadsheet.
        Returns (is_valid, error_message)."""
        headers_lower = [h.lower().strip() if h else "" for h in headers]
        required_lower = [r.lower().strip() for r in self.REQUIRED_COLUMNS]

        for req in required_lower:
            if req not in headers_lower:
                return False, f"Missing required column: '{req}'"
        return True, None

    def _column_index(self, headers: List[str], target: str) -> Optional[int]:
        """Find the column index of a header (case-insensitive)."""
        target_lower = target.lower().strip()
        for idx, h in enumerate(headers):
            if h and h.lower().strip() == target_lower:
                return idx
        return None

    def load(self) -> Tuple[bool, str]:
        """Load mappings from the spreadsheet.
        Returns (success, message) tuple."""

        if self._loaded:
            return True, f"Mappings already loaded ({len(self._mappings)} entries)"

        if not self._file_path:
            msg = "Bluebeam_Term_Mapping_1.xlsx not found in expected locations"
            logger.warning(msg)
            return False, msg

        if not os.path.exists(self._file_path):
            msg = f"Spreadsheet file not found: {self._file_path}"
            logger.warning(msg)
            return False, msg

        try:
            wb = load_workbook(self._file_path, read_only=True, data_only=True)
        except Exception as e:
            msg = f"Failed to open spreadsheet: {e}"
            logger.error(msg)
            return False, msg

        try:
            # Use the first/active sheet
            ws = wb.active
            if not ws:
                msg = "Spreadsheet has no active sheet"
                logger.error(msg)
                return False, msg

            # Read all rows
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                msg = "Spreadsheet is empty"
                logger.error(msg)
                return False, msg

            # First row should be headers
            headers = rows[0]
            if not headers:
                msg = "Spreadsheet has no headers"
                logger.error(msg)
                return False, msg

            # Validate required columns
            is_valid, err = self._validate_columns(headers)
            if not is_valid:
                msg = f"Spreadsheet validation failed: {err}"
                logger.error(msg)
                return False, msg

            # Find column indices
            desc_idx = self._column_index(headers, "Measurement Description")
            code_idx = self._column_index(headers, "Trade Code")
            group_idx = self._column_index(headers, "Trade Group")
            div_idx = self._column_index(headers, "CSI Division")
            unit_idx = self._column_index(headers, "Unit")
            type_idx = self._column_index(headers, "Takeoff Type")

            # Load data rows (skip header)
            self._mappings = []
            self._lookup_index = {}

            for row_num, row in enumerate(rows[1:], start=2):
                if not row:
                    continue

                # Extract values, handling None and empty strings
                description = self._safe_get(row, desc_idx, "").strip()
                if not description:
                    continue  # Skip rows with no description

                trade_code = self._safe_get(row, code_idx, "").strip()
                trade_group = self._safe_get(row, group_idx, "").strip()
                csi_division = self._safe_get(row, div_idx, "").strip()
                unit = self._safe_get(row, unit_idx, "").strip()
                takeoff_type = self._safe_get(row, type_idx, "").strip()

                mapping = {
                    "description": description,
                    "trade_code": trade_code,
                    "trade_group": trade_group,
                    "csi_division": csi_division,
                    "unit": unit,
                    "takeoff_type": takeoff_type,
                }
                self._mappings.append(mapping)

                # Index for fast lookups (lowercase)
                lookup_key = description.lower().strip()
                self._lookup_index[lookup_key] = mapping

            wb.close()
            self._loaded = True

            count = len(self._mappings)
            msg = f"Successfully loaded {count} measurement mappings"
            logger.info(msg)
            return True, msg

        except Exception as e:
            msg = f"Error loading mappings: {e}"
            logger.error(msg)
            return False, msg

    @staticmethod
    def _safe_get(row: tuple, idx: Optional[int], default: str = "") -> str:
        """Safely extract a value from a row tuple, converting to string."""
        if idx is None or idx >= len(row):
            return default
        val = row[idx]
        if val is None:
            return default
        return str(val).strip()

    def lookup(self, description: str) -> Optional[Dict]:
        """Look up a measurement description in the mappings.

        Tries (in order):
        1. Exact match (case-sensitive)
        2. Case-insensitive match

        Returns the mapping dict or None if not found.
        """
        if not self._loaded or not description:
            return None

        # Exact match first
        exact = description.strip()
        if exact in self._lookup_index:
            return self._lookup_index[exact]

        # Case-insensitive match
        lookup_key = exact.lower()
        if lookup_key in self._lookup_index:
            return self._lookup_index[lookup_key]

        return None

    def all_mappings(self) -> List[Dict]:
        """Return a list of all loaded mappings."""
        return self._mappings.copy()

    def status(self) -> Dict:
        """Return status of the loader."""
        return {
            "loaded": self._loaded,
            "count": len(self._mappings),
            "file_path": str(self._file_path) if self._file_path else None,
        }


# Global instance
_loader = MappingLoader()


def load_mappings_at_startup() -> Tuple[bool, str]:
    """Load mappings at application startup. Call once from app.py."""
    return _loader.load()


def lookup_measurement(description: str) -> Optional[Dict]:
    """Lookup a measurement description in the cached mappings.
    Returns the mapping dict or None if not found."""
    return _loader.lookup(description)


def get_all_mappings() -> List[Dict]:
    """Get all loaded mappings."""
    return _loader.all_mappings()


def get_loader_status() -> Dict:
    """Get the status of the mapping loader."""
    return _loader.status()
