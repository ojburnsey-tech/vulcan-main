# api/test_mapping_loader.py
# Unit tests for the Bluebeam Term Mapping loader.
#
# Tests:
# - Spreadsheet loads successfully
# - Required columns are validated
# - Number of mappings loaded is correct
# - Exact match works
# - Case-insensitive match works
# - Missing terms return None
# - Whitespace is trimmed

import unittest
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook
from mapping_loader import MappingLoader


class TestMappingLoader(unittest.TestCase):
    """Test suite for MappingLoader."""

    @classmethod
    def setUpClass(cls):
        """Create a temporary test spreadsheet."""
        cls.temp_dir = tempfile.mkdtemp()
        cls.test_file = Path(cls.temp_dir) / "Bluebeam_Term_Mapping_1.xlsx"

        # Create test workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Mappings"

        # Headers
        headers = [
            "Measurement Description",
            "Trade Code",
            "Trade Group",
            "CSI Division",
            "Unit",
            "Takeoff Type",
        ]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Test data
        test_data = [
            ["Ceramic floor tile", "TILE-CERAMIC", "Finishes", "09 3000", "m²", "Flooring"],
            ["Brick cavity wall", "BRICK-CAV", "Masonry", "04 4000", "m²", "Brickwork"],
            ["Concrete foundations", "CONC-FDN", "Structural", "03 3000", "m³", "Concreting"],
            ["Roof tiles pitched roof", "TILE-PITCH", "Roofing", "07 7000", "m²", "Roofing"],
            ["Emulsion paint to walls", "PAINT-EMULSION", "Finishes", "09 9100", "m²", "Decoration"],
        ]

        for row_num, row_data in enumerate(test_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        wb.save(cls.test_file)

    @classmethod
    def tearDownClass(cls):
        """Clean up temporary files."""
        import shutil
        shutil.rmtree(cls.temp_dir, ignore_errors=True)

    def test_loader_finds_spreadsheet(self):
        """Test that the loader can locate the spreadsheet."""
        loader = MappingLoader()
        # Mock the _find_mapping_file to return our test file
        loader._file_path = self.test_file
        ok, msg = loader.load()
        self.assertTrue(ok, f"Loading failed: {msg}")

    def test_loader_loads_correct_count(self):
        """Test that all mappings are loaded."""
        loader = MappingLoader()
        loader._file_path = self.test_file
        ok, msg = loader.load()
        self.assertTrue(ok)
        self.assertEqual(len(loader._mappings), 5, f"Expected 5 mappings, got {len(loader._mappings)}")

    def test_exact_match(self):
        """Test exact match lookup."""
        loader = MappingLoader()
        loader._file_path = self.test_file
        loader.load()

        result = loader.lookup("Ceramic floor tile")
        self.assertIsNotNone(result)
        self.assertEqual(result["trade_code"], "TILE-CERAMIC")
        self.assertEqual(result["trade_group"], "Finishes")
        self.assertEqual(result["csi_division"], "09 3000")
        self.assertEqual(result["unit"], "m²")
        self.assertEqual(result["takeoff_type"], "Flooring")

    def test_case_insensitive_match(self):
        """Test case-insensitive lookup."""
        loader = MappingLoader()
        loader._file_path = self.test_file
        loader.load()

        # Try different cases
        result1 = loader.lookup("CERAMIC FLOOR TILE")
        self.assertIsNotNone(result1)
        self.assertEqual(result1["trade_code"], "TILE-CERAMIC")

        result2 = loader.lookup("ceramic floor tile")
        self.assertIsNotNone(result2)
        self.assertEqual(result2["trade_code"], "TILE-CERAMIC")

        result3 = loader.lookup("CeRaMiC FlOoR TiLe")
        self.assertIsNotNone(result3)
        self.assertEqual(result3["trade_code"], "TILE-CERAMIC")

    def test_whitespace_trimming(self):
        """Test that leading/trailing whitespace is trimmed."""
        loader = MappingLoader()
        loader._file_path = self.test_file
        loader.load()

        result = loader.lookup("  Ceramic floor tile  ")
        self.assertIsNotNone(result)
        self.assertEqual(result["trade_code"], "TILE-CERAMIC")

        result = loader.lookup("\tBrick cavity wall\n")
        self.assertIsNotNone(result)
        self.assertEqual(result["trade_code"], "BRICK-CAV")

    def test_missing_term_returns_none(self):
        """Test that non-existent terms return None."""
        loader = MappingLoader()
        loader._file_path = self.test_file
        loader.load()

        result = loader.lookup("Nonexistent measurement")
        self.assertIsNone(result)

        result = loader.lookup("")
        self.assertIsNone(result)

        result = loader.lookup(None)
        self.assertIsNone(result)

    def test_all_mappings(self):
        """Test retrieving all mappings."""
        loader = MappingLoader()
        loader._file_path = self.test_file
        loader.load()

        all_mappings = loader.all_mappings()
        self.assertEqual(len(all_mappings), 5)
        self.assertIsInstance(all_mappings, list)

    def test_status(self):
        """Test loader status."""
        loader = MappingLoader()
        loader._file_path = self.test_file
        loader.load()

        status = loader.status()
        self.assertTrue(status["loaded"])
        self.assertEqual(status["count"], 5)
        self.assertIsNotNone(status["file_path"])

    def test_missing_file_returns_false(self):
        """Test that missing file is handled gracefully."""
        loader = MappingLoader()
        loader._file_path = Path(self.temp_dir) / "nonexistent.xlsx"
        ok, msg = loader.load()
        self.assertFalse(ok)
        self.assertIn("not found", msg.lower())

    def test_double_load_returns_cached(self):
        """Test that calling load twice returns cached result."""
        loader = MappingLoader()
        loader._file_path = self.test_file
        ok1, msg1 = loader.load()
        ok2, msg2 = loader.load()

        self.assertTrue(ok1)
        self.assertTrue(ok2)
        # Second load should indicate already loaded
        self.assertIn("already loaded", msg2.lower())


class TestMappingLoaderValidation(unittest.TestCase):
    """Test column validation."""

    def test_validate_columns_with_all_required(self):
        """Test validation passes with all required columns."""
        loader = MappingLoader()
        headers = [
            "Measurement Description",
            "Trade Code",
            "Trade Group",
            "CSI Division",
            "Unit",
            "Takeoff Type",
        ]
        is_valid, err = loader._validate_columns(headers)
        self.assertTrue(is_valid)
        self.assertIsNone(err)

    def test_validate_columns_case_insensitive(self):
        """Test validation is case-insensitive."""
        loader = MappingLoader()
        headers = [
            "measurement description",
            "TRADE CODE",
            "Trade Group",
            "csi division",
            "Unit",
            "takeoff type",
        ]
        is_valid, err = loader._validate_columns(headers)
        self.assertTrue(is_valid)
        self.assertIsNone(err)

    def test_validate_columns_missing_column(self):
        """Test validation fails with missing column."""
        loader = MappingLoader()
        headers = [
            "Measurement Description",
            "Trade Code",
            # Missing "Trade Group"
            "CSI Division",
            "Unit",
            "Takeoff Type",
        ]
        is_valid, err = loader._validate_columns(headers)
        self.assertFalse(is_valid)
        self.assertIn("missing", err.lower())


if __name__ == "__main__":
    unittest.main()
