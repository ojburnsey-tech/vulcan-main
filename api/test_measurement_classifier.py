# api/test_measurement_classifier.py
# Unit tests for the measurement classifier service.
#
# Tests:
# - Exact match classification
# - Case-insensitive classification
# - Batch classification
# - No match returns {matched: false}
# - Invalid input handling

import unittest
import tempfile
from pathlib import Path
from openpyxl import Workbook
from mapping_loader import MappingLoader, _loader
from measurement_classifier import MeasurementClassifier, classify_measurement


class TestMeasurementClassifier(unittest.TestCase):
    """Test suite for MeasurementClassifier."""

    @classmethod
    def setUpClass(cls):
        """Set up test environment with sample mappings."""
        # Create a temporary test spreadsheet
        cls.temp_dir = tempfile.mkdtemp()
        cls.test_file = Path(cls.temp_dir) / "Bluebeam_Term_Mapping_1.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Mappings"

        headers = [
            "Measurement Description",
            "Trade Code",
            "Trade Group",
            "CSI Division",
            "Unit",
            "Takeoff Type",
        ]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        test_data = [
            ["Ceramic floor tile", "TILE-CERAMIC", "Finishes", "09 3000", "m²", "Flooring"],
            ["Facing brick cavity wall", "BRICK-CAV", "Masonry", "04 4000", "m²", "Brickwork"],
            ["Reinforced concrete foundations", "CONC-FDN", "Structural", "03 3000", "m³", "Concreting"],
            ["Roof tiles pitched roof", "TILE-PITCH", "Roofing", "07 7000", "m²", "Roofing"],
            ["Emulsion paint to walls", "PAINT-EMULSION", "Finishes", "09 9100", "m²", "Decoration"],
        ]

        for row_num, row_data in enumerate(test_data, 2):
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        wb.save(cls.test_file)

        # Load mappings
        _loader._file_path = cls.test_file
        _loader.load()

    @classmethod
    def tearDownClass(cls):
        """Clean up."""
        import shutil
        shutil.rmtree(cls.temp_dir, ignore_errors=True)

    def test_classify_exact_match(self):
        """Test classification with exact match."""
        result = MeasurementClassifier.classify("Ceramic floor tile")

        self.assertTrue(result["matched"])
        self.assertEqual(result["description"], "Ceramic floor tile")
        self.assertEqual(result["trade_code"], "TILE-CERAMIC")
        self.assertEqual(result["trade_group"], "Finishes")
        self.assertEqual(result["csi_division"], "09 3000")
        self.assertEqual(result["unit"], "m²")
        self.assertEqual(result["takeoff_type"], "Flooring")

    def test_classify_case_insensitive(self):
        """Test classification is case-insensitive."""
        result = MeasurementClassifier.classify("CERAMIC FLOOR TILE")

        self.assertTrue(result["matched"])
        self.assertEqual(result["trade_code"], "TILE-CERAMIC")

    def test_classify_with_whitespace(self):
        """Test classification with whitespace."""
        result = MeasurementClassifier.classify("  Facing brick cavity wall  ")

        self.assertTrue(result["matched"])
        self.assertEqual(result["trade_code"], "BRICK-CAV")
        # Description should be trimmed
        self.assertEqual(result["description"], "Facing brick cavity wall")

    def test_classify_no_match(self):
        """Test classification returns matched=false for unknown descriptions."""
        result = MeasurementClassifier.classify("Unknown measurement type")

        self.assertFalse(result["matched"])
        self.assertEqual(result["description"], "Unknown measurement type")
        self.assertNotIn("trade_code", result)
        self.assertNotIn("trade_group", result)

    def test_classify_empty_description(self):
        """Test classification with empty description."""
        result = MeasurementClassifier.classify("")
        self.assertFalse(result["matched"])

        result = MeasurementClassifier.classify("   ")
        self.assertFalse(result["matched"])

    def test_classify_none_description(self):
        """Test classification with None description."""
        result = MeasurementClassifier.classify(None)
        self.assertFalse(result["matched"])

    def test_batch_classify(self):
        """Test batch classification."""
        descriptions = [
            "Ceramic floor tile",
            "Facing brick cavity wall",
            "Unknown item",
            "Emulsion paint to walls",
        ]

        results = MeasurementClassifier.batch_classify(descriptions)

        self.assertEqual(len(results), 4)
        self.assertTrue(results[0]["matched"])
        self.assertEqual(results[0]["trade_code"], "TILE-CERAMIC")
        self.assertTrue(results[1]["matched"])
        self.assertEqual(results[1]["trade_code"], "BRICK-CAV")
        self.assertFalse(results[2]["matched"])
        self.assertTrue(results[3]["matched"])
        self.assertEqual(results[3]["trade_code"], "PAINT-EMULSION")

    def test_batch_classify_empty_list(self):
        """Test batch classification with empty list."""
        results = MeasurementClassifier.batch_classify([])
        self.assertEqual(len(results), 0)

    def test_batch_classify_invalid_input(self):
        """Test batch classification with invalid input."""
        results = MeasurementClassifier.batch_classify("not a list")
        self.assertEqual(len(results), 0)

        results = MeasurementClassifier.batch_classify(None)
        self.assertEqual(len(results), 0)

    def test_classify_measurement_convenience_function(self):
        """Test the convenience function."""
        result = classify_measurement("Ceramic floor tile")

        self.assertTrue(result["matched"])
        self.assertEqual(result["trade_code"], "TILE-CERAMIC")

    def test_classify_all_test_cases(self):
        """Test classification of all test data."""
        test_cases = [
            ("Ceramic floor tile", "TILE-CERAMIC", "Finishes"),
            ("Facing brick cavity wall", "BRICK-CAV", "Masonry"),
            ("Reinforced concrete foundations", "CONC-FDN", "Structural"),
            ("Roof tiles pitched roof", "TILE-PITCH", "Roofing"),
            ("Emulsion paint to walls", "PAINT-EMULSION", "Finishes"),
        ]

        for description, expected_code, expected_group in test_cases:
            result = MeasurementClassifier.classify(description)
            self.assertTrue(result["matched"], f"Failed to match: {description}")
            self.assertEqual(result["trade_code"], expected_code)
            self.assertEqual(result["trade_group"], expected_group)

    def test_response_structure_matched(self):
        """Test that matched response has all required fields."""
        result = MeasurementClassifier.classify("Ceramic floor tile")

        required_fields = [
            "matched",
            "description",
            "trade_code",
            "trade_group",
            "csi_division",
            "unit",
            "takeoff_type",
        ]
        for field in required_fields:
            self.assertIn(field, result, f"Missing field: {field}")

    def test_response_structure_not_matched(self):
        """Test that unmatched response has only required fields."""
        result = MeasurementClassifier.classify("Unknown item")

        self.assertIn("matched", result)
        self.assertIn("description", result)
        self.assertFalse(result["matched"])
        # Should NOT have classification fields
        self.assertNotIn("trade_code", result)
        self.assertNotIn("trade_group", result)


if __name__ == "__main__":
    unittest.main()
