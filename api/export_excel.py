# api/export_excel.py
# Builds a fully-formatted NRM2-structured Excel BoQ from the enriched JSON
# that /process returns.  Called by the /export-excel route in app.py.
#
# The output is designed to drop straight into a QS firm's workflow:
#   - NRM2 trade sections with subtotals
#   - Formula cells so the QS can adjust rates and quantities live
#   - Rate split columns (material / labour) visible and editable
#   - "AI Measurement Draft" header — legally correct, professionally presented
#   - Firm name / project name written at the top for branding

import io
from openpyxl import Workbook
from totals import _effective_line_total
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# Shared logo decoder (handles data: URLs, http(s) URLs and local paths, and
# returns None on any problem) — same source the PDF letterhead uses, so both
# exports render the identical configured logo.
from export_pdf import _logo_image_bytes


# ── Colour palette (matches Vulcan Quanta brand) ──────────────────────────────
NAVY        = "0A1628"   # dark navy — trade section headers
GOLD        = "C8A96E"   # gold — title bar accent
MID_GREY    = "4A5568"   # column headers
LIGHT_GREY  = "F7F8FA"   # alternating row background
WHITE       = "FFFFFF"
AMBER_LIGHT = "FFFBEB"   # flagged item highlight
RED_LIGHT   = "FEF2F2"   # zero-rate warning


def _fill(hex_colour):
    """Return a solid PatternFill for a given hex colour string."""
    return PatternFill("solid", fgColor=hex_colour)


def _font(bold=False, colour="000000", size=10, italic=False):
    return Font(bold=bold, color=colour, size=size, italic=italic,
                name="Calibri")


def _border(style="thin"):
    s = Side(style=style, color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def _centre():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _right():
    return Alignment(horizontal="right", vertical="center")


# ── NRM2 column layout ────────────────────────────────────────────────────────
# Col A: Item ref  (narrow)
# Col B: Description (wide)
# Col C: Unit
# Col D: Quantity          ← QS edits this
# Col E: Material Rate     ← QS edits this
# Col F: Labour Rate       ← QS edits this
# Col G: Rate              pre-computed value (all 4 components, or flat rate fallback)
# Col H: Total (=D*G)      formula — auto-updates when QS edits D

COL_WIDTHS = {
    "A": 8,    # Item ref
    "B": 48,   # Description
    "C": 8,    # Unit
    "D": 12,   # Qty
    "E": 14,   # Mat Rate
    "F": 14,   # Lab Rate
    "G": 14,   # Rate
    "H": 16,   # Total
}

GBP = '#,##0.00'   # number format for currency cells


def _normalise_boq(boq_data):
    """
    Accept the same variety of JSON shapes that app.py's _enrich_boq handles.
    Always returns a list of  {trade: str, items: list}  dicts.
    """
    if isinstance(boq_data, list):
        return boq_data
    if isinstance(boq_data, dict):
        if "bill_of_quantities" in boq_data:
            return boq_data["bill_of_quantities"]
        if "trades" in boq_data:
            return boq_data["trades"]
        # Keys are trade names, values are item lists
        return [{"trade": k, "items": v}
                for k, v in boq_data.items() if isinstance(v, list)]
    return []


_LOGO_MAX_H_PX = 56    # cap the embedded logo height so it sits inside the info box
_LOGO_MAX_W_PX = 150   # cap width so a wide logo can't crowd the NRM2 labels


def _build_logo_xl_image(logo_src):
    """Return a scaled openpyxl Image for the branding logo, or None.

    Never raises — a broken or unsupported logo (e.g. SVG, which PIL cannot
    rasterise) must never break the Excel export.
    """
    raw = _logo_image_bytes(logo_src) if logo_src else None
    if not raw:
        return None
    try:
        from openpyxl.drawing.image import Image as XLImage   # needs Pillow
        img = XLImage(io.BytesIO(raw))
        if not img.width or not img.height:
            return None
        scale = min(_LOGO_MAX_H_PX / img.height, _LOGO_MAX_W_PX / img.width, 1.0)
        img.width  = int(img.width * scale)
        img.height = int(img.height * scale)
        return img
    except Exception:
        return None


def generate_boq_excel(boq_data, firm_name="", project_name="", branding=None):
    """
    Build the Excel workbook and return the raw bytes.

    Parameters
    ----------
    boq_data     : dict or list — the enriched BoQ JSON from /process
    firm_name    : str — the QS firm name written into the header (optional)
    project_name : str — the project name written into the header (optional)
    branding     : dict — the user's white-label branding (company_name,
                   company_address, company_phone, company_email, logo_url);
                   same shape the PDF pipeline receives (optional)

    Returns
    -------
    bytes — the .xlsx file content, ready to send as an HTTP response
    """
    brand = branding if isinstance(branding, dict) else {}

    groups = _normalise_boq(boq_data)
    if not groups:
        raise ValueError("No trade groups found in BoQ data.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Bill of Quantities"

    # ── Set column widths ─────────────────────────────────────────────────────
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    row = 1   # current row pointer — incremented as we write

    # ── Title block ───────────────────────────────────────────────────────────
    # Row 1: "VULCAN QUANTA" label
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = "VULCAN QUANTA — AI MEASUREMENT DRAFT"
    ws[f"A{row}"].fill      = _fill(NAVY)
    ws[f"A{row}"].font      = _font(bold=True, colour=GOLD, size=12)
    ws[f"A{row}"].alignment = _centre()
    ws.row_dimensions[row].height = 28
    row += 1

    # Row 2: disclaimer sub-line
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = (
        "AI-generated measurement draft — professional review and sign-off required "
        "before issue to client"
    )
    ws[f"A{row}"].fill      = _fill("1E3A5F")   # slightly lighter navy
    ws[f"A{row}"].font      = _font(italic=True, colour="94A3B8", size=9)
    ws[f"A{row}"].alignment = _centre()
    ws.row_dimensions[row].height = 18
    row += 1

    # Row 3: blank spacer
    ws.row_dimensions[row].height = 8
    row += 1

    # Firm / project info box. Branding contact lines appear only when
    # configured, so an unbranded export keeps the original four rows.
    labels = [
        ("Firm:",    firm_name or brand.get("company_name") or "[ Your firm name ]"),
        ("Project:", project_name or "[ Project name ]"),
    ]
    if brand.get("company_address"):
        labels.append(("Address:", brand["company_address"]))
    if brand.get("company_phone"):
        labels.append(("Phone:", brand["company_phone"]))
    if brand.get("company_email"):
        labels.append(("Email:", brand["company_email"]))
    labels += [
        ("Date:",    ""),   # left blank for QS to fill
        ("Prepared by:", "[ Name ]  [ MRICS / FRICS ]"),
    ]

    # Company logo floats over the empty E–F cells beside the info box.
    logo_img = _build_logo_xl_image(brand.get("logo_url") or brand.get("logo"))
    if logo_img is not None:
        ws.add_image(logo_img, f"E{row}")

    for label, value in labels:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font      = _font(bold=True, size=9, colour="374151")
        ws[f"A{row}"].alignment = _left()
        ws.merge_cells(f"B{row}:D{row}")
        ws[f"B{row}"] = value
        ws[f"B{row}"].font      = _font(size=9, colour="111827")
        ws[f"B{row}"].alignment = _left()
        # Right side: NRM2 label
        ws[f"G{row}"] = "NRM2" if label == "Firm:" else ""
        ws[f"H{row}"] = "Second Edition" if label == "Firm:" else ""
        ws[f"G{row}"].font = _font(bold=True, size=8, colour=MID_GREY)
        ws[f"H{row}"].font = _font(size=8, colour=MID_GREY)
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1   # blank gap before table

    # ── Column headers ────────────────────────────────────────────────────────
    header_labels = [
        ("A", "Ref"),
        ("B", "Description"),
        ("C", "Unit"),
        ("D", "Qty"),
        ("E", "Mat. Rate (£)"),
        ("F", "Lab. Rate (£)"),
        ("G", "Rate (£)"),
        ("H", "Total (£)"),
    ]
    header_row = row
    for col_letter, label in header_labels:
        cell = ws[f"{col_letter}{row}"]
        cell.value     = label
        cell.fill      = _fill(MID_GREY)
        cell.font      = _font(bold=True, colour=WHITE, size=9)
        cell.alignment = _centre()
        cell.border    = _border()
    ws.row_dimensions[row].height = 20
    row += 1

    # ── Trade sections ────────────────────────────────────────────────────────
    # Track the row numbers of every Total cell in column H so we can build the grand total
    grand_total_cells = []
    item_counter = 0   # sequential item reference number

    for group in groups:
        trade = group.get("trade", "Unspecified Trade")
        items = group.get("items") or group.get("line_items") or []
        if not isinstance(items, list):
            continue

        # Trade section header
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = trade.upper()
        ws[f"A{row}"].fill      = _fill(NAVY)
        ws[f"A{row}"].font      = _font(bold=True, colour=WHITE, size=10)
        ws[f"A{row}"].alignment = _left()
        ws[f"A{row}"].border    = _border()
        ws.row_dimensions[row].height = 20
        row += 1

        # Track row range for this trade's subtotal formula
        trade_item_rows = []
        shade = False   # alternating row colour toggle

        for item in items:
            if not isinstance(item, dict):
                continue

            item_counter += 1
            desc     = item.get("description") or item.get("desc") or ""
            unit     = item.get("unit", "")
            qty      = item.get("quantity") or item.get("qty") or 0
            mat_rate  = float(item.get("material_rate",       0) or 0)
            lab_rate  = float(item.get("labour_rate",         0) or 0)
            plant_rate = float(item.get("plant_rate",         0) or 0)
            waste_rate = float(item.get("waste_disposal_rate",0) or 0)
            comp_sum  = mat_rate + lab_rate + plant_rate + waste_rate
            flat_rate = float(item.get("rate", 0) or 0)
            eff_rate  = comp_sum if comp_sum > 0 else flat_rate

            bg = LIGHT_GREY if shade else WHITE
            shade = not shade

            # Flag items with zero rate so QS notices immediately
            if eff_rate == 0:
                bg = AMBER_LIGHT

            # Col A — item reference
            ws[f"A{row}"] = item_counter
            ws[f"A{row}"].fill      = _fill(bg)
            ws[f"A{row}"].font      = _font(size=9, colour="6B7280")
            ws[f"A{row}"].alignment = _centre()
            ws[f"A{row}"].border    = _border()

            # Col B — description
            ws[f"B{row}"] = desc
            ws[f"B{row}"].fill      = _fill(bg)
            ws[f"B{row}"].font      = _font(size=9)
            ws[f"B{row}"].alignment = _left()
            ws[f"B{row}"].border    = _border()

            # Col C — unit
            ws[f"C{row}"] = unit
            ws[f"C{row}"].fill      = _fill(bg)
            ws[f"C{row}"].font      = _font(size=9)
            ws[f"C{row}"].alignment = _centre()
            ws[f"C{row}"].border    = _border()

            # Col D — quantity (QS can edit this)
            ws[f"D{row}"] = qty
            ws[f"D{row}"].fill           = _fill(bg)
            ws[f"D{row}"].font           = _font(size=9)
            ws[f"D{row}"].alignment      = _right()
            ws[f"D{row}"].border         = _border()
            ws[f"D{row}"].number_format  = "0.00"

            # Col E — material rate (QS can edit this)
            ws[f"E{row}"] = mat_rate
            ws[f"E{row}"].fill           = _fill(bg)
            ws[f"E{row}"].font           = _font(size=9)
            ws[f"E{row}"].alignment      = _right()
            ws[f"E{row}"].border         = _border()
            ws[f"E{row}"].number_format  = GBP

            # Col F — labour rate (QS can edit this)
            ws[f"F{row}"] = lab_rate
            ws[f"F{row}"].fill           = _fill(bg)
            ws[f"F{row}"].font           = _font(size=9)
            ws[f"F{row}"].alignment      = _right()
            ws[f"F{row}"].border         = _border()
            ws[f"F{row}"].number_format  = GBP

            # Col G — effective rate as a pre-computed value (all 4 components, or flat rate)
            ws[f"G{row}"] = eff_rate
            ws[f"G{row}"].fill           = _fill(bg)
            ws[f"G{row}"].font           = _font(size=9)
            ws[f"G{row}"].alignment      = _right()
            ws[f"G{row}"].border         = _border()
            ws[f"G{row}"].number_format  = GBP

            # Col H — line total formula: =D*G
            ws[f"H{row}"] = f"=D{row}*G{row}"
            ws[f"H{row}"].fill           = _fill(bg)
            ws[f"H{row}"].font           = _font(size=9, colour="1D4ED8")  # blue = formula cell
            ws[f"H{row}"].alignment      = _right()
            ws[f"H{row}"].border         = _border()
            ws[f"H{row}"].number_format  = GBP

            trade_item_rows.append(row)
            ws.row_dimensions[row].height = 16
            row += 1

        # Trade subtotal row
        if trade_item_rows:
            first_item = trade_item_rows[0]
            last_item  = trade_item_rows[-1]

            ws[f"A{row}"] = ""
            ws.merge_cells(f"B{row}:G{row}")
            ws[f"B{row}"] = f"{trade.title()} — Subtotal"
            ws[f"B{row}"].fill      = _fill("E8F4FD")
            ws[f"B{row}"].font      = _font(bold=True, size=9, colour=NAVY)
            ws[f"B{row}"].alignment = _right()
            ws[f"B{row}"].border    = _border()

            # Subtotal formula sums the H column for this trade's items
            ws[f"H{row}"] = f"=SUM(H{first_item}:H{last_item})"
            ws[f"H{row}"].fill           = _fill("DBEAFE")
            ws[f"H{row}"].font           = _font(bold=True, size=9, colour=NAVY)
            ws[f"H{row}"].alignment      = _right()
            ws[f"H{row}"].border         = _border()
            ws[f"H{row}"].number_format  = GBP
            ws.row_dimensions[row].height = 18

            grand_total_cells.append(f"H{row}")
            row += 1

        row += 1   # blank gap between trades

    # ── Grand total ───────────────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"] = "TOTAL (excl. VAT)"
    ws[f"A{row}"].fill      = _fill(NAVY)
    ws[f"A{row}"].font      = _font(bold=True, colour=GOLD, size=11)
    ws[f"A{row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"A{row}"].border    = _border()

    # Grand total = sum of all trade subtotals
    grand_formula = "=" + "+".join(grand_total_cells) if grand_total_cells else "=0"
    ws[f"H{row}"] = grand_formula
    ws[f"H{row}"].fill          = _fill(NAVY)
    ws[f"H{row}"].font          = _font(bold=True, colour=GOLD, size=11)
    ws[f"H{row}"].alignment     = _right()
    ws[f"H{row}"].border        = _border()
    ws[f"H{row}"].number_format = GBP
    ws.row_dimensions[row].height = 24
    row += 2

    # ── Footer disclaimer ─────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:H{row}")
    ws[f"A{row}"] = (
        "AI Measurement Draft prepared by Vulcan Quanta. "
        "Reviewed and approved by: ______________________  "
        "MRICS / FRICS  |  Date: ____________  |  "
        "Rates based on BCIS Q2 2026. Verify before tender issue."
    )
    ws[f"A{row}"].font      = _font(italic=True, size=8, colour="9CA3AF")
    ws[f"A{row}"].alignment = _left()
    ws.row_dimensions[row].height = 30

    # ── Sign-off certificate (only written when the BoQ has been signed off) ──
    signoff = boq_data.get('signoff') if isinstance(boq_data, dict) else None
    if isinstance(signoff, dict) and signoff.get('signed_off_by'):
        import datetime as _dt
        row += 3

        # Section heading
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = "SIGN-OFF CERTIFICATE"
        ws[f"A{row}"].fill      = _fill(NAVY)
        ws[f"A{row}"].font      = _font(bold=True, colour=GOLD, size=11)
        ws[f"A{row}"].alignment = _centre()
        ws.row_dimensions[row].height = 22
        row += 1

        name     = str(signoff.get("signed_off_by") or "")
        title    = str(signoff.get("signoff_title")  or "")
        raw_at   = signoff.get("signed_off_at") or ""
        h        = str(signoff.get("signoff_hash") or "")

        try:
            dt = _dt.datetime.fromisoformat(raw_at.replace("Z", "+00:00"))
            signed_at_str = dt.strftime("%-d %B %Y at %H:%M UTC")
        except Exception:
            signed_at_str = raw_at

        cert_lines = [
            ("Signed off by",  name),
            ("Qualification",  title),
            ("Date & time",    signed_at_str),
            ("Integrity hash", f"SHA-256: {h[:16]}…{h[-8:]}"),
        ]
        for lbl, val in cert_lines:
            ws[f"A{row}"] = lbl
            ws[f"A{row}"].font      = _font(bold=True, size=9, colour="374151")
            ws[f"A{row}"].alignment = _left()
            ws.merge_cells(f"B{row}:H{row}")
            ws[f"B{row}"] = val
            ws[f"B{row}"].font      = _font(size=9, colour="111827")
            ws[f"B{row}"].alignment = _left()
            ws.row_dimensions[row].height = 16
            row += 1

    # ── Freeze panes so headers stay visible when scrolling ───────────────────
    ws.freeze_panes = ws[f"A{header_row + 1}"]

    # ── Write to bytes in memory — never touches the filesystem ───────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()